"""Gercek sozlesme tarifelerinin ice aktarilmasi -- **gercek veri yuvasi**.

Motorun sentetik veriyle calisan hali ile gercek Ozdilek sozlesmeleriyle calisan
hali arasindaki tek fark bu modul. Cekirdek motorun (`tariff`, `packing`, `risk`,
`sla`, `decision`) hicbir satiri degismeden gercek fiyatlara gecilebilir.

Kargo firmalari sozlesme tarifelerini genellikle **matris** halinde gonderir:
satirlar desi kademeleri, sutunlar bolgeler. Bu bicim `data/carriers/*.yaml`
semasina cevrilir ve ceviri sirasinda dogrulanir.

Iki giris bicimi destekleniyor:

`from_csv`
    Bagimlilik gerektirmez. Excel'den "CSV olarak kaydet" ile uretilebilir;
    pratikte en dayanikli yol.

`from_excel`
    `openpyxl` gerektirir (`pip install -e ".[excel]"`). Firmadan gelen dosyayi
    elle donusturmeye gerek birakmaz.

**En onemli davranis:** ice aktarilan tarife `source: contract` isaretlenir.
Bu bayrak arayuze kadar gider; "ORNEK TARIFE" rozeti ancak o zaman kaybolur.
Sentetik bir fiyatin gercek sozlesme fiyati sanilmasi, bu projedeki en ciddi
yanlis anlasilma riski.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import yaml

from ..domain.enums import CarrierCode, RoundingRule, TariffSourceKind, ZoneClass
from ..tariff.schema import Tariff

#: Matris CSV'sinde beklenen ilk sutun.
DESI_COLUMN = "up_to_desi"

#: Bolge sutunlarinin basliklari (`ZoneClass` degerleriyle ayni olmali).
ZONE_COLUMNS = tuple(zone.value for zone in ZoneClass)


class ContractImportError(ValueError):
    """Sozlesme dosyasi okunamadi veya dogrulamadan gecemedi."""


@dataclass(frozen=True, slots=True)
class ContractMeta:
    """Fiyat matrisinde bulunmayan, sozlesmenin geri kalanini tanimlayan alanlar.

    Bunlar firmadan gelen matris dosyasinda genellikle yoktur; sozlesme metninden
    veya satin alma ekibinden alinir. Varsayilan verilmedi -- sessizce yanlis bir
    yakit farki varsaymaktansa acikca sorulmasi daha iyi.
    """

    carrier: CarrierCode
    display_name: str
    valid_from: date
    min_charge: float
    rounding: RoundingRule
    fuel_pct: float
    cod_fee: float
    insurance_free_limit: float
    insurance_pct_above: float
    vat_pct: float
    over_top_tier_per_desi: dict[str, float]
    sla_days: dict[str, int]
    cutoff: str
    max_desi_per_parcel: float
    rural_extra_days: int = 1
    cod_supported: bool = True
    unserved_plates: tuple[int, ...] = ()
    volume_discounts: tuple[tuple[int, float], ...] = ()
    note: str = ""


def _read_matrix_rows(rows: list[dict[str, Any]], source_name: str) -> list[dict[str, Any]]:
    """Matris satirlarini desi kademelerine cevirir ve dogrular."""
    if not rows:
        raise ContractImportError(f"{source_name}: fiyat matrisi bos")

    headers = set(rows[0])
    missing = {DESI_COLUMN, *ZONE_COLUMNS} - headers
    if missing:
        raise ContractImportError(
            f"{source_name}: eksik sutun {sorted(missing)}. "
            f"Beklenen basliklar: {DESI_COLUMN}, {', '.join(ZONE_COLUMNS)}"
        )

    tiers: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):  # 1 = baslik satiri
        try:
            up_to = float(row[DESI_COLUMN])
            zones = {zone: float(row[zone]) for zone in ZONE_COLUMNS}
        except (TypeError, ValueError) as exc:
            raise ContractImportError(
                f"{source_name}: {index}. satirda sayiya cevrilemeyen deger -- {exc}"
            ) from exc

        if up_to <= 0:
            raise ContractImportError(f"{source_name}: {index}. satirda desi pozitif olmali")
        tiers.append({"up_to": up_to, "zones": zones})

    return sorted(tiers, key=lambda tier: tier["up_to"])


def build_tariff(tiers: list[dict[str, Any]], meta: ContractMeta) -> Tariff:
    """Matris + sozlesme alanlarindan dogrulanmis bir `Tariff` uretir.

    Dogrulama `Tariff` semasinda yapilir: kademelerin siralanmis olmasi, her
    kademede tum bolgelerin bulunmasi ve **fiyatlarin desi arttikca dusmemesi**.
    Sonuncusu, elle duzenlenmis bir matriste en sik goruleni ve fark edilmesi en
    zor olan hatadir.
    """
    payload = {
        "carrier": meta.carrier.value,
        "display_name": meta.display_name,
        "source": TariffSourceKind.CONTRACT.value,
        "note": meta.note,
        "valid_from": meta.valid_from.isoformat(),
        "currency": "TRY",
        "rounding": meta.rounding.value,
        "desi_step": 1.0,
        "min_charge": meta.min_charge,
        "desi_tiers": tiers,
        "over_30_per_desi": meta.over_top_tier_per_desi,
        "surcharges": {
            "fuel_pct": meta.fuel_pct,
            "cod_fee": meta.cod_fee,
            "insurance": {
                "free_limit": meta.insurance_free_limit,
                "pct_above": meta.insurance_pct_above,
            },
            "vat_pct": meta.vat_pct,
        },
        "volume_discounts": [
            {"monthly_parcels_gte": threshold, "pct": pct}
            for threshold, pct in meta.volume_discounts
        ],
        "service": {
            "sla_days": meta.sla_days,
            "rural_extra_days": meta.rural_extra_days,
            "cutoff": meta.cutoff,
        },
        "constraints": {
            "max_desi_per_parcel": meta.max_desi_per_parcel,
            "cod_supported": meta.cod_supported,
            "unserved_plates": list(meta.unserved_plates),
        },
    }

    try:
        return Tariff.model_validate(payload)
    except Exception as exc:
        raise ContractImportError(
            f"{meta.carrier.value}: sozlesme tarifesi dogrulamadan gecemedi --\n{exc}"
        ) from exc


def from_csv(path: Path, meta: ContractMeta) -> Tariff:
    """Matris CSV'sinden sozlesme tarifesi okur.

    Beklenen bicim:

        up_to_desi,sehir_ici,bolge_ici,bolgeler_arasi,uzak
        1,68.00,76.00,88.00,99.00
        2,74.00,83.00,96.00,108.00
        ...
    """
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    except OSError as exc:
        raise ContractImportError(f"{path.name}: dosya okunamadi -- {exc}") from exc

    return build_tariff(_read_matrix_rows(rows, path.name), meta)


def from_excel(path: Path, meta: ContractMeta, sheet: str | None = None) -> Tariff:
    """Matris Excel dosyasindan sozlesme tarifesi okur.

    `openpyxl` gerektirir. Eksikse hata mesaji nasil kurulacagini soyler --
    "ModuleNotFoundError: openpyxl" tek basina bir kullaniciya bir sey anlatmaz.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - istege bagli bagimlilik
        raise ContractImportError(
            "Excel ice aktarimi icin `openpyxl` gerekli. Kurulum:\n"
            '    pip install -e ".[excel]"\n'
            "Alternatif olarak dosyayi CSV kaydedip `from_csv` kullanabilirsiniz."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    if worksheet is None:
        raise ContractImportError(f"{path.name}: sayfa bulunamadi")

    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(iterator)]
    except StopIteration as exc:
        raise ContractImportError(f"{path.name}: dosya bos") from exc

    rows = [
        dict(zip(headers, values, strict=False))
        for values in iterator
        if any(value is not None for value in values)
    ]
    workbook.close()

    return build_tariff(_read_matrix_rows(rows, path.name), meta)


def write_yaml(tariff: Tariff, directory: Path) -> Path:
    """Ice aktarilmis tarifeyi motorun okudugu dizine yazar.

    Boylece ice aktarim bir kez yapilir; motor bundan sonra her zamanki gibi
    `data/carriers/*.yaml` dosyalarini okur. Ayri bir "gercek veri modu" kodu yok.
    """
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"{tariff.carrier.value.lower()}.yaml"

    payload = tariff.model_dump(mode="json")
    header = (
        f"# {tariff.display_name} -- GERCEK SOZLESME TARIFESI\n"
        f"# Ice aktarim: desi_engine.adapters.contract_import\n"
        f"# Gecerlilik baslangici: {tariff.valid_from}\n"
    )
    with path.open("w", encoding="utf-8") as handle:
        handle.write(header)
        yaml.safe_dump(payload, handle, allow_unicode=True, sort_keys=False, width=100)
    return path
